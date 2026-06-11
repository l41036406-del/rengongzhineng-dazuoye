from __future__ import annotations

import json
import os
import re
import unicodedata
from io import BytesIO
from datetime import datetime
from difflib import get_close_matches
from functools import lru_cache
from pathlib import Path
from typing import Literal
from xml.sax.saxutils import escape

import joblib
import numpy as np
import pandas as pd
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
MODEL_DIR = BASE_DIR / "models"
OUTPUT_DIR = BASE_DIR / "outputs"
FIGURE_DIR = OUTPUT_DIR / "figures"
WEB_DIST = BASE_DIR / "web" / "dist"

LABELS = ["home_win", "draw", "away_win"]
LABEL_NAMES = {"home_win": "队伍A胜", "draw": "平局", "away_win": "队伍B胜"}
FORMATION_PROFILES = {
    "4-3-3": {
        "attack": 0.76,
        "control": 0.60,
        "defense": 0.54,
        "press": 0.74,
        "width": 0.82,
        "description": "强调边路宽度与前场压迫，进攻上限较高。",
    },
    "4-2-3-1": {
        "attack": 0.65,
        "control": 0.70,
        "defense": 0.69,
        "press": 0.63,
        "width": 0.66,
        "description": "双后腰保护中路，攻守转换更稳定。",
    },
    "4-4-2": {
        "attack": 0.58,
        "control": 0.53,
        "defense": 0.64,
        "press": 0.54,
        "width": 0.70,
        "description": "结构紧凑、双前锋直接，整体均衡。",
    },
    "3-5-2": {
        "attack": 0.67,
        "control": 0.79,
        "defense": 0.57,
        "press": 0.59,
        "width": 0.64,
        "description": "中场人数占优，依赖翼卫覆盖两侧空间。",
    },
    "5-3-2": {
        "attack": 0.43,
        "control": 0.55,
        "defense": 0.85,
        "press": 0.39,
        "width": 0.47,
        "description": "低位防守稳定，主要通过反击制造机会。",
    },
    "3-4-3": {
        "attack": 0.83,
        "control": 0.56,
        "defense": 0.43,
        "press": 0.77,
        "width": 0.80,
        "description": "前场投入积极，但身后空间风险更高。",
    },
    "4-1-4-1": {
        "attack": 0.49,
        "control": 0.71,
        "defense": 0.76,
        "press": 0.59,
        "width": 0.57,
        "description": "单后腰连接两线，适合控制节奏与封锁中路。",
    },
}


class MatchRequest(BaseModel):
    home_team: str = Field(min_length=1)
    away_team: str = Field(min_length=1)
    neutral: bool = True
    is_world_cup_final: bool = True
    model_name: str | None = None
    home_formation: str = "4-2-3-1"
    away_formation: str = "4-3-3"
    date: str | None = None
    tournament: str | None = None


class BatchRequest(BaseModel):
    matches: list[MatchRequest] = Field(min_length=1, max_length=500)


app = FastAPI(
    title="World Cup Intelligence API",
    version="1.0.0",
    docs_url="/api/docs",
    openapi_url="/api/openapi.json",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        os.getenv("CORS_ORIGIN", "https://yfs.onrender.com"),
        "https://worldcup-predict.azurewebsites.net",
        "http://worldcup-predict.azurewebsites.net",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if FIGURE_DIR.exists():
    app.mount("/figures", StaticFiles(directory=FIGURE_DIR), name="figures")


@lru_cache
def clean_data() -> pd.DataFrame:
    return pd.read_csv(DATA_DIR / "world_cup_clean.csv", parse_dates=["date"])


@lru_cache
def feature_data() -> pd.DataFrame:
    return pd.read_csv(DATA_DIR / "world_cup_features.csv", parse_dates=["date"])


@lru_cache
def team_data() -> dict:
    with (DATA_DIR / "team_state.json").open(encoding="utf-8") as file:
        return json.load(file)


TEAM_ALIASES = {
    "us": "United States",
    "usa": "United States",
    "united states of america": "United States",
    "美国": "United States",
    "eng": "England",
    "英格兰": "England",
    "china pr": "China",
    "pr china": "China",
    "中国": "China",
    "korea republic": "South Korea",
    "republic of korea": "South Korea",
    "korea rep": "South Korea",
    "韩国": "South Korea",
    "korea dpr": "North Korea",
    "dpr korea": "North Korea",
    "朝鲜": "North Korea",
    "ir iran": "Iran",
    "islamic republic of iran": "Iran",
    "cote d ivoire": "Ivory Coast",
    "côte d ivoire": "Ivory Coast",
    "cabo verde": "Cape Verde",
    "congo dr": "DR Congo",
    "congo kinshasa": "DR Congo",
    "democratic republic of congo": "DR Congo",
    "democratic republic of the congo": "DR Congo",
    "republic of the congo": "Congo",
    "congo brazzaville": "Congo",
    "czechia": "Czech Republic",
    "turkiye": "Turkey",
    "türkiye": "Turkey",
    "fyr macedonia": "North Macedonia",
    "macedonia": "North Macedonia",
    "ireland": "Republic of Ireland",
    "eire": "Republic of Ireland",
    "uae": "United Arab Emirates",
    "hong kong china": "Hong Kong",
    "chinese taipei": "Taiwan",
    "viet nam": "Vietnam",
    "kyrgyz republic": "Kyrgyzstan",
    "republic of moldova": "Moldova",
    "russian federation": "Russia",
    "brunei darussalam": "Brunei",
    "lao pdr": "Laos",
    "syrian arab republic": "Syria",
    "state of palestine": "Palestine",
    "swaziland": "Eswatini",
    "burma": "Myanmar",
    "east timor": "Timor-Leste",
    "bosnia herzgovina": "Bosnia and Herzegovina",
    "bosnia herzegovina": "Bosnia and Herzegovina",
    "trinidad tobago": "Trinidad and Tobago",
    "antigua barbuda": "Antigua and Barbuda",
    "st kitts and nevis": "Saint Kitts and Nevis",
    "st lucia": "Saint Lucia",
    "st vincent and the grenadines": "Saint Vincent and the Grenadines",
    "sao tome and principe": "São Tomé and Príncipe",
}


def team_name_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value.strip().casefold())
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", normalized)


@lru_cache
def team_name_lookup() -> dict[str, str]:
    names = list(team_data()["team_state"])
    lookup = {team_name_key(name): name for name in names}
    for name in names:
        lookup.setdefault(team_name_key(name.replace(" and ", " & ")), name)
        if name.startswith("Saint "):
            lookup.setdefault(team_name_key(name.replace("Saint ", "St ", 1)), name)
    for alias, canonical in TEAM_ALIASES.items():
        if canonical in names:
            lookup[team_name_key(alias)] = canonical
    return lookup


def canonical_team_name(value: str) -> str:
    raw = value.strip()
    canonical = team_name_lookup().get(team_name_key(raw))
    if canonical:
        return canonical
    candidates = get_close_matches(raw, list(team_data()["team_state"]), n=3, cutoff=0.58)
    suggestion = f"；可能是：{'、'.join(candidates)}" if candidates else ""
    raise HTTPException(
        status_code=422,
        detail=f"无法识别队伍“{raw}”{suggestion}。请使用标准队名或常见国家别名。",
    )


def normalize_match_request(request: MatchRequest) -> MatchRequest:
    for formation in (request.home_formation, request.away_formation):
        if formation not in FORMATION_PROFILES:
            raise HTTPException(status_code=422, detail=f"暂不支持阵型“{formation}”")
    return request.model_copy(
        update={
            "home_team": canonical_team_name(request.home_team),
            "away_team": canonical_team_name(request.away_team),
        }
    )


@lru_cache
def evaluation_data() -> dict:
    with (OUTPUT_DIR / "eval_results.json").open(encoding="utf-8") as file:
        return json.load(file)


@lru_cache
def model_bundle():
    return {
        "best": joblib.load(MODEL_DIR / "best_model.pkl"),
        "scaler": joblib.load(MODEL_DIR / "scaler.pkl"),
        "models": joblib.load(MODEL_DIR / "all_models.pkl"),
        "features": joblib.load(MODEL_DIR / "feature_columns.pkl"),
    }


def build_feature_vector(request: MatchRequest) -> dict[str, float]:
    state = team_data()
    teams = state["team_state"]
    h2h = state["h2h"]
    year = state["data_end_year"]
    default = {
        "elo": 1500.0,
        "recent_winrate": 0.5,
        "recent_goal_diff": 0.0,
        "avg_goals": 1.0,
        "wc_exp": 0,
    }
    home = teams.get(request.home_team, default)
    away = teams.get(request.away_team, default)
    home_h2h = h2h.get(request.home_team, {}).get(request.away_team, 0)
    away_h2h = h2h.get(request.away_team, {}).get(request.home_team, 0)
    return {
        "home_elo": home["elo"],
        "away_elo": away["elo"],
        "elo_diff": home["elo"] - away["elo"],
        "elo_abs_diff": home["elo"] + away["elo"],
        "home_recent_winrate": home["recent_winrate"],
        "away_recent_winrate": away["recent_winrate"],
        "home_recent_goal_diff": home["recent_goal_diff"],
        "away_recent_goal_diff": away["recent_goal_diff"],
        "home_avg_goals": home["avg_goals"],
        "away_avg_goals": away["avg_goals"],
        "h2h_diff": home_h2h - away_h2h,
        "home_wc_exp": home["wc_exp"],
        "away_wc_exp": away["wc_exp"],
        "neutral": int(request.neutral),
        "is_world_cup_final": int(request.is_world_cup_final),
        "match_year": int(year),
    }


def model_probability_map(model, scaled: np.ndarray) -> dict[str, float]:
    probabilities = {label: 0.0 for label in LABELS}
    if hasattr(model, "predict_proba"):
        predicted = model.predict_proba(scaled)[0]
        classes = list(model.classes_)
        for label in LABELS:
            if label in classes:
                probabilities[label] = float(predicted[classes.index(label)])
    else:
        probabilities[str(model.predict(scaled)[0])] = 1.0
    return probabilities


def elo_prior(vector: dict[str, float]) -> dict[str, float]:
    difference = float(vector["elo_diff"])
    expected_home = 1.0 / (1.0 + 10 ** (-difference / 400.0))
    draw_probability = 0.08 + 0.20 * np.exp(-abs(difference) / 250.0)
    decisive_probability = 1.0 - draw_probability
    return {
        "home_win": float(decisive_probability * expected_home),
        "draw": float(draw_probability),
        "away_win": float(decisive_probability * (1.0 - expected_home)),
    }


def consensus_probabilities(bundle: dict, scaled: np.ndarray, vector: dict[str, float]):
    metrics = evaluation_data()["metrics"]
    totals = {label: 0.0 for label in LABELS}
    weight_total = 0.0
    votes = {LABEL_NAMES[label]: 0 for label in LABELS}
    details = []
    for name, model in bundle["models"].items():
        model_probabilities = model_probability_map(model, scaled)
        metric = metrics.get(name, {})
        weight = float(
            metric.get(
                "composite",
                (metric.get("accuracy", 0.5) + metric.get("macro_f1", 0.5)) / 2,
            )
        )
        weight = max(weight, 0.1)
        for label in LABELS:
            totals[label] += model_probabilities[label] * weight
        weight_total += weight
        winner = max(LABELS, key=model_probabilities.get)
        votes[LABEL_NAMES[winner]] += 1
        details.append(
            {
                "model": name,
                "prediction": LABEL_NAMES[winner],
                "confidence": round(model_probabilities[winner] * 100, 1),
            }
        )

    ensemble = {label: totals[label] / weight_total for label in LABELS}
    prior = elo_prior(vector)
    blended = {
        label: 0.82 * ensemble[label] + 0.18 * prior[label] for label in LABELS
    }
    return blended, votes, details, prior


def apply_formation_adjustment(
    probabilities: dict[str, float],
    home_formation: str,
    away_formation: str,
):
    home = FORMATION_PROFILES[home_formation]
    away = FORMATION_PROFILES[away_formation]
    home_edge = (
        0.52 * (home["attack"] - away["defense"])
        + 0.24 * (home["control"] - away["control"])
        + 0.16 * (home["press"] - away["control"])
        + 0.08 * (home["width"] - away["width"])
    )
    away_edge = (
        0.52 * (away["attack"] - home["defense"])
        + 0.24 * (away["control"] - home["control"])
        + 0.16 * (away["press"] - home["control"])
        + 0.08 * (away["width"] - home["width"])
    )
    draw_edge = (
        0.50 * (((home["defense"] + away["defense"]) / 2) - 0.60)
        + 0.32 * (0.22 - abs(home["control"] - away["control"]))
    )
    # Increase formation impact enough to be decision-relevant, but keep a hard cap
    # so tactical setup does not overwhelm team strength and model consensus.
    home_shift = float(np.clip(home_edge * 0.95, -0.60, 0.60))
    draw_shift = float(np.clip(draw_edge * 0.70, -0.42, 0.42))
    away_shift = float(np.clip(away_edge * 0.95, -0.60, 0.60))
    adjustments = np.array([home_shift, draw_shift, away_shift])
    base = np.array([max(probabilities[label], 1e-6) for label in LABELS])
    logits = np.log(base) + adjustments
    adjusted = np.exp(logits - logits.max())
    adjusted = adjusted / adjusted.sum()
    result = {label: float(adjusted[index]) for index, label in enumerate(LABELS)}
    deltas = {
        LABEL_NAMES[label]: round((result[label] - probabilities[label]) * 100, 1)
        for label in LABELS
    }
    strongest = max(deltas, key=lambda label: abs(deltas[label]))
    direction = "提升" if deltas[strongest] >= 0 else "降低"
    return result, {
        "home_formation": home_formation,
        "away_formation": away_formation,
        "home_description": home["description"],
        "away_description": away["description"],
        "impact_score": round(max(abs(home_shift), abs(draw_shift), abs(away_shift)), 2),
        "probability_delta": deltas,
        "summary": f"阵型情景对“{strongest}”影响最大，概率{direction} {abs(deltas[strongest]):.1f} 个百分点。",
        "method": "阵型影响为独立战术情景修正，不属于历史训练特征；本次已提高权重，但仍受硬上限约束。",
    }


def run_prediction(request: MatchRequest) -> dict:
    request = normalize_match_request(request)
    if request.home_team == request.away_team:
        raise HTTPException(status_code=400, detail="请选择两支不同的球队")

    bundle = model_bundle()
    model = bundle["best"]
    if request.model_name:
        model = bundle["models"].get(request.model_name)
        if model is None:
            raise HTTPException(status_code=400, detail="未找到指定模型")

    vector = build_feature_vector(request)
    features = bundle["features"]
    values = np.array([[vector[column] for column in features]], dtype=float)
    scaled = bundle["scaler"].transform(values)
    model_votes = {}
    model_details = []
    prior = elo_prior(vector)
    if request.model_name:
        raw_probabilities = model_probability_map(model, scaled)
        winner = max(LABELS, key=raw_probabilities.get)
        model_votes = {LABEL_NAMES[label]: int(label == winner) for label in LABELS}
        model_details = [
            {
                "model": request.model_name,
                "prediction": LABEL_NAMES[winner],
                "confidence": round(raw_probabilities[winner] * 100, 1),
            }
        ]
        model_label = request.model_name
    else:
        raw_probabilities, model_votes, model_details, prior = consensus_probabilities(
            bundle, scaled, vector
        )
        model_label = "多模型共识 + ELO 校准"

    adjusted_probabilities, formation_impact = apply_formation_adjustment(
        raw_probabilities, request.home_formation, request.away_formation
    )
    prediction = max(LABELS, key=adjusted_probabilities.get)
    probabilities = {
        LABEL_NAMES[label]: adjusted_probabilities[label] for label in LABELS
    }
    base_probabilities = {
        LABEL_NAMES[label]: raw_probabilities[label] for label in LABELS
    }

    states = team_data()["team_state"]
    warnings = [
        team for team in (request.home_team, request.away_team) if team not in states
    ]
    return {
        "home_team": request.home_team,
        "away_team": request.away_team,
        "date": request.date,
        "tournament": request.tournament,
        "neutral": request.neutral,
        "is_world_cup_final": request.is_world_cup_final,
        "home_formation": request.home_formation,
        "away_formation": request.away_formation,
        "prediction": LABEL_NAMES.get(prediction, prediction),
        "probabilities": probabilities,
        "base_probabilities": base_probabilities,
        "home_elo": states.get(request.home_team, {}).get("elo", 1500),
        "away_elo": states.get(request.away_team, {}).get("elo", 1500),
        "model": model_label,
        "model_votes": model_votes,
        "model_details": model_details,
        "elo_prior": {LABEL_NAMES[label]: prior[label] for label in LABELS},
        "formation_impact": formation_impact,
        "feature_snapshot": {
            "elo_diff": round(vector["elo_diff"], 1),
            "home_recent_winrate": round(vector["home_recent_winrate"] * 100, 1),
            "away_recent_winrate": round(vector["away_recent_winrate"] * 100, 1),
            "home_recent_goal_diff": round(vector["home_recent_goal_diff"], 2),
            "away_recent_goal_diff": round(vector["away_recent_goal_diff"], 2),
        },
        "warnings": warnings,
    }


def build_batch_xlsx(matches: list[MatchRequest]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    results = [run_prediction(match) for match in matches]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "预测结果"
    headers = [
        "比赛日期",
        "赛事名称",
        "队伍A",
        "队伍B",
        "预测结果",
        "队伍A胜概率",
        "平局概率",
        "队伍B胜概率",
        "队伍A ELO",
        "队伍B ELO",
        "中立场地",
        "世界杯正赛",
        "队伍A阵型",
        "队伍B阵型",
        "使用模型",
    ]
    worksheet.append(headers)
    for item in results:
        worksheet.append(
            [
                item["date"] or "",
                item["tournament"] or "",
                item["home_team"],
                item["away_team"],
                item["prediction"],
                item["probabilities"]["队伍A胜"],
                item["probabilities"]["平局"],
                item["probabilities"]["队伍B胜"],
                item["home_elo"],
                item["away_elo"],
                "是" if item["neutral"] else "否",
                "是" if item["is_world_cup_final"] else "否",
                item["home_formation"],
                item["away_formation"],
                item["model"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="17372F")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
        for cell in row[5:8]:
            cell.number_format = "0.0%"

    widths = [13, 20, 20, 20, 12, 15, 13, 15, 13, 13, 12, 14, 13, 13, 24]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 24

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def team_history_summary(team: str) -> dict:
    data = clean_data()
    matches = data[(data["home_team"] == team) | (data["away_team"] == team)].copy()
    if matches.empty:
        return {
            "matches": 0,
            "wins": 0,
            "draws": 0,
            "losses": 0,
            "win_rate": 0.0,
            "goals_for": 0.0,
            "goals_against": 0.0,
        }

    matches = matches.sort_values("date", ascending=False).head(10)
    is_home = matches["home_team"] == team
    goals_for = np.where(is_home, matches["home_score"], matches["away_score"])
    goals_against = np.where(is_home, matches["away_score"], matches["home_score"])
    wins = int((goals_for > goals_against).sum())
    draws = int((goals_for == goals_against).sum())
    losses = int((goals_for < goals_against).sum())
    count = len(matches)
    return {
        "matches": count,
        "wins": wins,
        "draws": draws,
        "losses": losses,
        "win_rate": round(wins / count, 4),
        "goals_for": round(float(goals_for.mean()), 2),
        "goals_against": round(float(goals_against.mean()), 2),
    }


def team_recommendations(team: dict, fixtures: list[dict]) -> list[str]:
    recommendations = []
    strong_opponents = [item for item in fixtures if item["opponent_elo"] - team["elo"] >= 80]
    close_matches = [item for item in fixtures if item["confidence"] < 0.5]
    weaker_opponents = [item for item in fixtures if team["elo"] - item["opponent_elo"] >= 100]

    if team["recent_goal_diff"] < 0:
        recommendations.append(
            f"近期净胜球为 {team['recent_goal_diff']:.1f}，防线承压明显；面对高位压迫时应减少后场横传，并优先保护禁区弧顶。"
        )
    elif team["recent_goal_diff"] >= 1:
        recommendations.append(
            f"近期场均净胜球达到 {team['recent_goal_diff']:.1f}，进攻状态突出；建议延续主动压迫，在领先后降低无效对攻。"
        )

    if strong_opponents:
        names = "、".join(item["opponent"] for item in strong_opponents[:3])
        recommendations.append(
            f"对阵 {names} 时处于明显实力劣势，建议采用更紧凑的中低位防守，把进攻重点放在定位球和快速反击。"
        )
    if weaker_opponents:
        names = "、".join(item["opponent"] for item in weaker_opponents[:3])
        recommendations.append(
            f"对阵 {names} 时拥有较大 ELO 优势，应提高前场压迫强度，但需预留防守宽度以防对手反击。"
        )
    if close_matches:
        names = "、".join(item["opponent"] for item in close_matches[:3])
        recommendations.append(
            f"与 {names} 的预测分布较接近，临场阵容和场地因素可能改变结果，建议重点准备替补调整与最后 30 分钟策略。"
        )
    if team["recent_winrate"] < 40:
        recommendations.append(
            f"近期胜率仅 {team['recent_winrate']:.0f}%，应优先提升比赛稳定性，避免在均势阶段过早投入过多进攻人数。"
        )
    elif team["recent_winrate"] >= 70:
        recommendations.append(
            f"近期胜率为 {team['recent_winrate']:.0f}%，状态处于高位；轮换应保持中轴线连续性，避免破坏现有攻防节奏。"
        )
    return recommendations[:4]


def build_dynamic_report(matches: list[MatchRequest]) -> dict:
    predictions = [run_prediction(match) for match in matches]
    states = team_data()["team_state"]
    team_names = sorted(
        {name for item in predictions for name in (item["home_team"], item["away_team"])}
    )
    team_reports = []

    for name in team_names:
        state = states.get(
            name,
            {
                "elo": 1500.0,
                "recent_winrate": 0.5,
                "recent_goal_diff": 0.0,
                "avg_goals": 1.0,
                "wc_exp": 0,
                "games_played": 0,
            },
        )
        fixtures = []
        expected_wins = expected_draws = expected_losses = expected_points = 0.0
        for index, item in enumerate(predictions, start=1):
            if name not in (item["home_team"], item["away_team"]):
                continue
            is_home = item["home_team"] == name
            win_probability = item["probabilities"]["队伍A胜" if is_home else "队伍B胜"]
            draw_probability = item["probabilities"]["平局"]
            loss_probability = item["probabilities"]["队伍B胜" if is_home else "队伍A胜"]
            opponent = item["away_team"] if is_home else item["home_team"]
            opponent_elo = float(states.get(opponent, {}).get("elo", 1500))
            confidence = max(win_probability, draw_probability, loss_probability)
            expected_wins += win_probability
            expected_draws += draw_probability
            expected_losses += loss_probability
            expected_points += 3 * win_probability + draw_probability
            fixtures.append(
                {
                    "match": index,
                    "date": item["date"],
                    "opponent": opponent,
                    "side": "队伍A" if is_home else "队伍B",
                    "formation": (
                        item["home_formation"] if is_home else item["away_formation"]
                    ),
                    "opponent_formation": (
                        item["away_formation"] if is_home else item["home_formation"]
                    ),
                    "opponent_elo": round(opponent_elo, 1),
                    "win": round(win_probability * 100, 1),
                    "draw": round(draw_probability * 100, 1),
                    "loss": round(loss_probability * 100, 1),
                    "confidence": round(confidence, 4),
                    "prediction": (
                        "胜"
                        if item["prediction"] == ("队伍A胜" if is_home else "队伍B胜")
                        else "负"
                        if item["prediction"] == ("队伍B胜" if is_home else "队伍A胜")
                        else "平"
                    ),
                }
            )

        history = team_history_summary(name)
        report = {
            "team": name,
            "elo": round(float(state.get("elo", 1500)), 1),
            "recent_winrate": round(float(state.get("recent_winrate", 0.5)) * 100, 1),
            "recent_goal_diff": round(float(state.get("recent_goal_diff", 0)), 2),
            "avg_goals": round(float(state.get("avg_goals", 1)), 2),
            "wc_exp": int(state.get("wc_exp", 0)),
            "fixtures_count": len(fixtures),
            "expected_wins": round(expected_wins, 2),
            "expected_draws": round(expected_draws, 2),
            "expected_losses": round(expected_losses, 2),
            "expected_points": round(expected_points, 2),
            "points_per_match": round(expected_points / len(fixtures), 2),
            "history": history,
            "fixtures": fixtures,
        }
        report["recommendations"] = team_recommendations(report, fixtures)
        team_reports.append(report)

    outcome_counts = {"队伍A胜": 0, "平局": 0, "队伍B胜": 0}
    for item in predictions:
        outcome_counts[item["prediction"]] += 1
        probabilities = item["probabilities"]
        ordered = sorted(probabilities.items(), key=lambda pair: pair[1], reverse=True)
        item["confidence"] = round(ordered[0][1] * 100, 1)
        item["probability_gap"] = round((ordered[0][1] - ordered[1][1]) * 100, 1)
        item["elo_gap"] = round(float(item["home_elo"]) - float(item["away_elo"]), 1)

    average_confidence = float(np.mean([item["confidence"] for item in predictions]))
    uncertain_matches = sum(item["confidence"] < 50 for item in predictions)
    generated_at = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S")
    summary = {
        "matches": len(predictions),
        "teams": len(team_names),
        "average_confidence": round(average_confidence, 1),
        "uncertain_matches": uncertain_matches,
        "strongest_team": max(team_reports, key=lambda item: item["elo"])["team"],
        "best_schedule_outlook": max(
            team_reports, key=lambda item: item["points_per_match"]
        )["team"],
    }

    markdown_lines = [
        "# 上传赛程实时分析报告",
        "",
        f"- 生成时间：{generated_at}",
        f"- 分析比赛：{summary['matches']} 场",
        f"- 涉及队伍：{summary['teams']} 支",
        f"- 平均预测置信度：{summary['average_confidence']}%",
        "",
        "## 总体判断",
        "",
        f"本批数据中，ELO 最高的队伍是 **{summary['strongest_team']}**；按每场预期积分计算，赛程前景最佳的是 **{summary['best_schedule_outlook']}**。"
        f"共有 **{summary['uncertain_matches']}** 场比赛的最高结果概率低于 50%，这些对局更容易受到首发、伤停和临场战术影响。",
        "",
        "## 队伍分析",
    ]
    for team in sorted(team_reports, key=lambda item: item["expected_points"], reverse=True):
        markdown_lines.extend(
            [
                "",
                f"### {team['team']}",
                "",
                f"- ELO：{team['elo']}；近期胜率：{team['recent_winrate']}%；近期净胜球：{team['recent_goal_diff']}",
                f"- {team['fixtures_count']} 场比赛预期积分：{team['expected_points']}，场均 {team['points_per_match']}",
                f"- 预期胜/平/负：{team['expected_wins']} / {team['expected_draws']} / {team['expected_losses']}",
                "",
            ]
        )
        markdown_lines.extend(f"- 建议：{text}" for text in team["recommendations"])

    return {
        "generated_at": generated_at,
        "summary": summary,
        "outcome_distribution": [
            {"name": name, "value": value} for name, value in outcome_counts.items()
        ],
        "matches": predictions,
        "teams": sorted(team_reports, key=lambda item: item["expected_points"], reverse=True),
        "markdown": "\n".join(markdown_lines),
    }


def build_report_pdf(report: dict) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font_name = "STSong-Light"
    font_candidates = [
        os.getenv("REPORT_FONT_PATH"),
        str(BASE_DIR / "assets" / "fonts" / "NotoSansCJKsc-Regular.otf"),
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    ]
    for font_path in font_candidates:
        if font_path and Path(font_path).exists():
            font_name = "ReportChinese"
            pdfmetrics.registerFont(TTFont(font_name, font_path, subfontIndex=0))
            break
    else:
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=17 * mm,
        bottomMargin=16 * mm,
        title="世界杯预测分析报告",
        author="世界杯比赛结果智能预测与分析系统",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=22,
        leading=30,
        textColor=colors.HexColor("#17372f"),
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    heading = ParagraphStyle(
        "ChineseHeading",
        parent=styles["Heading2"],
        fontName=font_name,
        fontSize=14,
        leading=20,
        textColor=colors.HexColor("#17372f"),
        spaceBefore=9,
        spaceAfter=7,
    )
    body = ParagraphStyle(
        "ChineseBody",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=9.5,
        leading=16,
        textColor=colors.HexColor("#40584f"),
        spaceAfter=5,
    )
    small = ParagraphStyle(
        "ChineseSmall",
        parent=body,
        fontSize=8,
        leading=12,
    )
    table_header = ParagraphStyle(
        "ChineseTableHeader",
        parent=small,
        textColor=colors.white,
        alignment=TA_CENTER,
        leading=11,
    )
    story = [
        Paragraph("世界杯预测分析报告", title),
        Paragraph(f"生成时间：{escape(report['generated_at'])}", body),
        Spacer(1, 4 * mm),
        Paragraph("总体摘要", heading),
    ]
    summary = report["summary"]
    summary_data = [
        ["分析比赛", "涉及队伍", "平均置信度", "高不确定对局"],
        [
            f"{summary['matches']} 场",
            f"{summary['teams']} 支",
            f"{summary['average_confidence']}%",
            f"{summary['uncertain_matches']} 场",
        ],
    ]
    summary_table = Table(summary_data, colWidths=[41 * mm] * 4)
    summary_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#52675f")),
                ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#17372f")),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, 1), 14),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#cbc2af")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#a9a18f")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbc2af")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8f3e8")),
            ]
        )
    )
    story.extend(
        [
            summary_table,
            Spacer(1, 5 * mm),
            Paragraph("总体判断", heading),
            Paragraph(
                f"ELO 最高的队伍是 {escape(summary['strongest_team'])}；"
                f"赛程前景最佳的是 {escape(summary['best_schedule_outlook'])}。"
                f"共有 {summary['uncertain_matches']} 场比赛的最高结果概率低于 50%，"
                "应重点关注首发、伤停与临场战术变化。",
                body,
            ),
            Paragraph("比赛预测明细", heading),
        ]
    )
    match_rows = [
        ["队伍A", "队伍B", "双方阵型", "预测", "队伍A胜", "平局", "队伍B胜", "置信度"]
    ]
    for item in report["matches"]:
        match_rows.append(
            [
                item["home_team"],
                item["away_team"],
                f"{item['home_formation']} / {item['away_formation']}",
                item["prediction"],
                f"{item['probabilities']['队伍A胜'] * 100:.1f}%",
                f"{item['probabilities']['平局'] * 100:.1f}%",
                f"{item['probabilities']['队伍B胜'] * 100:.1f}%",
                f"{item['confidence']:.1f}%",
            ]
        )
    match_table = Table(
        [
            [
                Paragraph(
                    escape(str(cell)),
                    table_header if row_index == 0 else small,
                )
                for cell in row
            ]
            for row_index, row in enumerate(match_rows)
        ],
        colWidths=[23 * mm, 23 * mm, 25 * mm, 19 * mm, 18 * mm, 15 * mm, 18 * mm, 19 * mm],
        repeatRows=1,
    )
    match_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17372f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("ALIGN", (2, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbc2af")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f3e8")]),
                ("TOPPADDING", (0, 0), (-1, 0), 7),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
                ("TOPPADDING", (0, 1), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
            ]
        )
    )
    story.append(match_table)

    for index, team in enumerate(report["teams"]):
        if index and index % 3 == 0:
            story.append(PageBreak())
        story.extend(
            [
                Paragraph(escape(team["team"]), heading),
                Paragraph(
                    f"ELO：{team['elo']}　近期胜率：{team['recent_winrate']}%　"
                    f"近期净胜球：{team['recent_goal_diff']}　场均进球：{team['avg_goals']}",
                    body,
                ),
                Paragraph(
                    f"{team['fixtures_count']} 场比赛预期积分：{team['expected_points']}，"
                    f"场均预期积分：{team['points_per_match']}；"
                    f"预期胜/平/负：{team['expected_wins']} / {team['expected_draws']} / {team['expected_losses']}",
                    body,
                ),
            ]
        )
        for recommendation in team["recommendations"]:
            story.append(Paragraph(f"• {escape(recommendation)}", body))

    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor("#718078"))
        canvas.drawCentredString(A4[0] / 2, 9 * mm, f"第 {doc.page} 页")
        canvas.restoreState()

    document.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)
    return buffer


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/api/overview")
def overview():
    data = clean_data()
    result_counts = (
        data["result"].map(LABEL_NAMES).value_counts().rename_axis("name").reset_index(name="value")
    )
    recent = data.sort_values("date", ascending=False).head(10).copy()
    recent["date"] = recent["date"].dt.strftime("%Y-%m-%d")
    recent["result"] = recent["result"].map(LABEL_NAMES)
    records = recent[
        ["date", "home_team", "away_team", "home_score", "away_score", "tournament", "result"]
    ].to_dict(orient="records")
    return {
        "metrics": {
            "matches": len(data),
            "finals": int(data["is_world_cup_final"].sum()),
            "qualifiers": int((1 - data["is_world_cup_final"]).sum()),
            "year_min": int(data["date"].dt.year.min()),
            "year_max": int(data["date"].dt.year.max()),
        },
        "recent_matches": records,
        "result_distribution": result_counts.to_dict(orient="records"),
        "fields": [
            {"field": "date", "meaning": "比赛日期", "type": "日期"},
            {"field": "home_team / away_team", "meaning": "记录中的对阵双方", "type": "类别"},
            {"field": "home_score / away_score", "meaning": "双方进球数", "type": "数值"},
            {"field": "tournament", "meaning": "世界杯正赛或预选赛", "type": "类别"},
            {"field": "neutral", "meaning": "是否为中立场地", "type": "布尔"},
            {"field": "result", "meaning": "队伍A胜 / 平局 / 队伍B胜", "type": "目标标签"},
        ],
    }


@app.get("/api/models")
def models():
    evaluation = evaluation_data()
    metrics = [
        {"name": name, **values}
        for name, values in evaluation["metrics"].items()
    ]
    tuning = [
        {"k": int(k), **values}
        for k, values in evaluation.get("knn_tuning", {}).items()
    ]
    return {
        "best_model": evaluation["best_model_name"],
        "best_macro_f1_model": evaluation["best_macro_f1_name"],
        "best_k": evaluation.get("best_k"),
        "metrics": metrics,
        "knn_tuning": sorted(tuning, key=lambda item: item["k"]),
        "model_names": list(model_bundle()["models"].keys()),
        "formations": [
            {"name": name, **profile} for name, profile in FORMATION_PROFILES.items()
        ],
    }


@app.get("/api/teams")
def teams():
    states = team_data()["team_state"]
    return {
        "teams": [
            {"name": name, **values}
            for name, values in sorted(states.items())
        ]
    }


@app.post("/api/predict")
def predict(request: MatchRequest):
    return run_prediction(request)


@app.post("/api/batch")
def batch_predict(request: BatchRequest):
    return {"results": [run_prediction(match) for match in request.matches]}


@app.post("/api/batch/export/xlsx")
def export_batch_xlsx(request: BatchRequest):
    filename = f"batch-predictions-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    workbook = build_batch_xlsx(request.matches)
    content = workbook.getvalue()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(content)),
        },
    )


@app.get("/api/report")
def report():
    path = OUTPUT_DIR / "reports" / "analysis_report.md"
    if not path.exists():
        raise HTTPException(status_code=404, detail="分析报告尚未生成")
    return {"markdown": path.read_text(encoding="utf-8")}


@app.post("/api/report/analyze")
def analyze_report(request: BatchRequest):
    return build_dynamic_report(request.matches)


@app.post("/api/report/export/pdf")
def export_report_pdf(request: BatchRequest):
    filename = f"world-cup-analysis-{datetime.now().strftime('%Y%m%d-%H%M%S')}.pdf"
    try:
        report_data = build_dynamic_report(request.matches)
        pdf_buffer = build_report_pdf(report_data)
        content = pdf_buffer.getvalue()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"PDF 报告生成失败：{exc}") from exc
    return Response(
        content=content,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(content)),
        },
    )


@app.get("/api/visualizations")
def visualizations():
    clean = clean_data().copy()
    features = feature_data().copy()
    evaluation = evaluation_data()

    clean["year"] = clean["date"].dt.year
    clean["total_goals"] = clean["home_score"] + clean["away_score"]
    trend = (
        clean.groupby("year")
        .agg(matches=("result", "size"), avg_goals=("total_goals", "mean"))
        .reset_index()
    )

    result_distribution = (
        clean["result"].value_counts().reindex(LABELS).fillna(0).astype(int)
    )

    goal_rows = []
    for goals in range(0, 9):
        goal_rows.append(
            {
                "goals": str(goals),
                "队伍A": int((clean["home_score"] == goals).sum()),
                "队伍B": int((clean["away_score"] == goals).sum()),
                "总进球": int((clean["total_goals"] == goals).sum()),
            }
        )
    goal_rows.append(
        {
            "goals": "9+",
            "队伍A": int((clean["home_score"] >= 9).sum()),
            "队伍B": int((clean["away_score"] >= 9).sum()),
            "总进球": int((clean["total_goals"] >= 9).sum()),
        }
    )

    team_records: dict[str, dict[str, int]] = {}
    for row in clean[["home_team", "away_team", "result"]].itertuples(index=False):
        team_records.setdefault(row.home_team, {"matches": 0, "wins": 0})
        team_records.setdefault(row.away_team, {"matches": 0, "wins": 0})
        team_records[row.home_team]["matches"] += 1
        team_records[row.away_team]["matches"] += 1
        if row.result == "home_win":
            team_records[row.home_team]["wins"] += 1
        elif row.result == "away_win":
            team_records[row.away_team]["wins"] += 1
    strong_teams = sorted(
        (
            {
                "team": team,
                "matches": values["matches"],
                "win_rate": round(values["wins"] / values["matches"] * 100, 1),
            }
            for team, values in team_records.items()
        ),
        key=lambda item: item["matches"],
        reverse=True,
    )[:12]
    strong_teams.sort(key=lambda item: item["win_rate"])

    venue_rows = []
    for neutral, label in ((False, "非中立场"), (True, "中立场")):
        subset = clean[clean["neutral"].astype(bool) == neutral]
        distribution = subset["result"].value_counts(normalize=True).reindex(LABELS).fillna(0)
        venue_rows.append(
            {
                "venue": label,
                **{
                    LABEL_NAMES[result]: round(float(distribution[result]) * 100, 1)
                    for result in LABELS
                },
            }
        )

    elo_edges = np.arange(1000, 2001, 50)
    home_counts, _ = np.histogram(features["home_elo"], bins=elo_edges)
    away_counts, _ = np.histogram(features["away_elo"], bins=elo_edges)
    elo_distribution = [
        {
            "elo": f"{int(elo_edges[index])}",
            "队伍A": int(home_counts[index]),
            "队伍B": int(away_counts[index]),
        }
        for index in range(len(home_counts))
    ]

    corr_columns = [
        "elo_diff",
        "home_recent_winrate",
        "away_recent_winrate",
        "home_recent_goal_diff",
        "away_recent_goal_diff",
        "home_avg_goals",
        "away_avg_goals",
        "h2h_diff",
        "neutral",
        "is_world_cup_final",
    ]
    corr = features[corr_columns].corr().round(2)
    correlations = [
        {"x": x, "y": y, "value": float(corr.loc[y, x])}
        for y in corr_columns
        for x in corr_columns
    ]

    model_metrics = [
        {
            "model": name,
            "accuracy": round(values["accuracy"] * 100, 2),
            "macro_f1": round(values["macro_f1"] * 100, 2),
        }
        for name, values in evaluation["metrics"].items()
    ]

    bundle = model_bundle()
    importance_model = bundle["best"]
    if not hasattr(importance_model, "feature_importances_"):
        importance_model = bundle["models"].get("随机森林")
    importances = getattr(importance_model, "feature_importances_", np.zeros(len(bundle["features"])))
    feature_importance = sorted(
        (
            {"feature": feature, "importance": round(float(value) * 100, 2)}
            for feature, value in zip(bundle["features"], importances)
        ),
        key=lambda item: item["importance"],
    )

    return {
        "trend": trend.round({"avg_goals": 2}).to_dict(orient="records"),
        "result_distribution": [
            {"name": LABEL_NAMES[name], "value": int(result_distribution[name])}
            for name in LABELS
        ],
        "goal_distribution": goal_rows,
        "strong_teams": strong_teams,
        "venue_comparison": venue_rows,
        "elo_distribution": elo_distribution,
        "correlation": {"columns": corr_columns, "values": correlations},
        "model_metrics": model_metrics,
        "feature_importance": feature_importance,
    }


if WEB_DIST.exists():
    app.mount("/assets", StaticFiles(directory=WEB_DIST / "assets"), name="web-assets")

    @app.get("/{path:path}", include_in_schema=False)
    def serve_spa(path: str):
        requested = WEB_DIST / path
        if path and requested.is_file():
            return FileResponse(requested)
        return FileResponse(WEB_DIST / "index.html")
